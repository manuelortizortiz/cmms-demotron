import os, re, math, json, unicodedata
from datetime import datetime, date
from functools import wraps
from io import BytesIO
from flask import Flask, request, redirect, url_for, session, render_template_string, jsonify, send_file
from sqlalchemy import create_engine, text, inspect
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY','demotron-secret')
DATABASE_URL = os.getenv('DATABASE_URL','').strip()
if DATABASE_URL.startswith('postgres://'):
    DATABASE_URL = DATABASE_URL.replace('postgres://','postgresql+psycopg://',1)
elif DATABASE_URL.startswith('postgresql://'):
    DATABASE_URL = DATABASE_URL.replace('postgresql://','postgresql+psycopg://',1)
engine = create_engine(DATABASE_URL, pool_pre_ping=True, future=True) if DATABASE_URL else None

ESTADOS=['AL DÍA','PRÓXIMA','ATRASADA','EN PROCESO','POR RECIBIR','EN TALLER','FUERA DE SERVICIO','PROGRAMADO']
TIPOS_PM=['PM1','PM2','PM3','PM4','PM5','CORRECTIVA']
UBICACIONES=['Palmucho','Quirihue','Curico','Taller Central','Villa Seca','Cobquecura','Pelluhue','San Carlos','San Nicolas','Linares','Talca','Taller Externo']

def dialect(): return engine.dialect.name if engine is not None else ''
def pk_sql(): return 'INTEGER PRIMARY KEY AUTOINCREMENT' if dialect()=='sqlite' else 'SERIAL PRIMARY KEY'
def safe(v):
    if v is None: return ''
    if isinstance(v,(datetime,date)): return v.strftime('%Y-%m-%d')
    s=str(v); return '' if s.lower() in ['nan','none','nat'] else s
def clean(v):
    if v is None: return None
    s=str(v).strip(); return None if not s or s.lower() in ['none','nan','nat'] else s
def up(v):
    s=clean(v); return s.upper() if s else None
def norm_key(v):
    v='' if v is None else str(v).strip().lower()
    v=''.join(c for c in unicodedata.normalize('NFKD',v) if not unicodedata.combining(c))
    v=re.sub(r'[^a-z0-9]+','_',v)
    return re.sub(r'_+','_',v).strip('_')
def norm_ubic(v):
    if v is None: return ''
    raw=str(v).strip()
    if raw.lower() in ['','nan','none','nat']: return ''
    k=norm_key(raw).replace('_',' ')
    mapa={'palmucho':'Palmucho','q61':'Palmucho','q 61':'Palmucho','quirihue':'Quirihue','curico':'Curico','taller':'Taller Central','taller central':'Taller Central','villa seca':'Villa Seca','villaseca':'Villa Seca','cobquecura':'Cobquecura','pelluhue':'Pelluhue','san carlos':'San Carlos','san nicolas':'San Nicolas','linares':'Linares','talca':'Talca','taller externo':'Taller Externo'}
    return mapa.get(k, raw.title())
def number(v):
    s=clean(v)
    if s is None: return None
    try:
        s=s.replace('$','').replace('CLP','').replace('clp','').replace(' ','')
        if ',' in s: s=s.replace('.','').replace(',','.')
        elif s.count('.')>1: s=s.replace('.','')
        return float(s)
    except: return None
def num(v):
    if v is None: return 0
    if isinstance(v,(int,float)):
        try:
            if isinstance(v,float) and math.isnan(v): return 0
        except: pass
        return float(v)
    return number(v) or 0
def clp(v): return '$ '+format(int(round(num(v))), ',').replace(',','.')
def q(sql, params=None, fetch=True):
    if engine is None: return []
    with engine.begin() as conn:
        res=conn.execute(text(sql), params or {})
        return [dict(r) for r in res.mappings().all()] if fetch else []
def tables():
    try: return inspect(engine).get_table_names() if engine else []
    except: return []
def table_exists(t): return t in tables()
def columns(t):
    try: return [c['name'] for c in inspect(engine).get_columns(t)]
    except: return []
def col_exists(t,c): return c in columns(t)
def add_col(t,c,ddl='TEXT'):
    if table_exists(t) and not col_exists(t,c):
        try: q(f'ALTER TABLE {t} ADD COLUMN {c} {ddl}', fetch=False)
        except: pass
def first(cols,*needles,exclude=()):
    for n in needles:
        for c in cols:
            lc=c.lower()
            if lc==n.lower() and all(x not in lc for x in exclude): return c
    for n in needles:
        for c in cols:
            lc=c.lower()
            if n.lower() in lc and all(x not in lc for x in exclude): return c
    return None
def val(r,c,d=''): return safe(r.get(c,d)) if c else d

def map_cols(t):
    cols=columns(t)
    if t in ['maestro_equipos','equipos']:
        return {'codigo':first(cols,'codigo','cod_equipo','equipo'),'tipo_equipo':first(cols,'tipo_equipo','tipo_de_equipo','tipo'),'familia':first(cols,'familia','grupo'),'marca':first(cols,'marca'),'modelo':first(cols,'modelo'),'ano':first(cols,'ano','anio','año'),'ubicacion':first(cols,'ubicacion','obra','faena','destino'),'responsable':first(cols,'responsable','operador'),'lectura_actual':first(cols,'lectura_actual','horometro','horas','kilometraje','odometro','km'),'unidad':first(cols,'unidad','control'),'proxima_pm':first(cols,'proxima_pm','proxima','proxima_mantencion'),'estado':first(cols,'estado_cmms','estado_servicio','estado_operativo','estado',exclude=('control',)),'descripcion':first(cols,'descripcion','detalle')}
    if t=='lecturas':
        return {'id':first(cols,'id'),'fecha':first(cols,'fecha','fecha_de_combustible'),'codigo':first(cols,'codigo','cod_equipo','equipo'),'horometro':first(cols,'horometro','horometros','horas'),'kilometraje':first(cols,'kilometraje','kilometro','odometro','km'),'obra_ubicacion':first(cols,'obra_ubicacion','ubicacion','obra','faena','destino'),'responsable':first(cols,'responsable','conductor','operador'),'observacion':first(cols,'observacion','comentario','nota')}
    if t=='mantenciones':
        return {'id':first(cols,'id'),'fecha':first(cols,'fecha'),'codigo':first(cols,'codigo','cod_equipo','equipo'),'tipo_mantencion':first(cols,'tipo_mantencion','tipo','pm'),'lectura':first(cols,'lectura','horometro','kilometraje'),'espm':first(cols,'espm','descripcion','detalle'),'folio':first(cols,'folio','ot','orden'),'lugar':first(cols,'lugar','ubicacion','obra'),'proveedor':first(cols,'proveedor','responsable'),'costo_mantencion_clp':first(cols,'costo_mantencion_clp','costo','monto','valor'),'estado':first(cols,'estado')}
    if t=='ot':
        return {'id':first(cols,'id'),'fecha':first(cols,'fecha'),'ot':first(cols,'ot','folio','orden'),'codigo':first(cols,'codigo','cod_equipo','equipo'),'tipo':first(cols,'tipo','tipo_mantencion','pm'),'lectura':first(cols,'lectura','horometro','kilometraje'),'descripcion':first(cols,'descripcion','detalle','trabajo','espm'),'responsable':first(cols,'responsable','proveedor'),'estado':first(cols,'estado'),'costo':first(cols,'costo','monto','valor')}
    if t=='compras':
        return {'id':first(cols,'id'),'fecha':first(cols,'fecha'),'oc':first(cols,'oc','orden_compra'),'codigo':first(cols,'codigo','equipo'),'descripcion':first(cols,'descripcion','detalle'),'proveedor':first(cols,'proveedor'),'costo_pm_clp':first(cols,'costo_pm_clp','monto','valor','total'),'estado_oc':first(cols,'estado_oc','estado')}
    if t=='bodega':
        return {'id':first(cols,'id'),'folio':first(cols,'folio'),'fecha':first(cols,'fecha'),'equipo':first(cols,'equipo'),'envio':first(cols,'envio'),'persona_que_retiro':first(cols,'persona_que_retiro','retira'),'destino':first(cols,'destino','ubicacion'),'comentario':first(cols,'comentario','observacion'),'codigo':first(cols,'codigo','cod_equipo')}
    if t in ['plan_mantenciones','plan_90_dias','planner_semanal']:
        return {'id':first(cols,'id'),'codigo':first(cols,'codigo','equipo'),'tipo_equipo':first(cols,'tipo_equipo','tipo'),'familia':first(cols,'familia'),'control':first(cols,'control','unidad'),'lectura_actual':first(cols,'lectura_actual','horometro','kilometraje'),'proxima_lectura_objetivo':first(cols,'proxima_lectura_objetivo','proxima'),'promedio_diario':first(cols,'promedio_diario','promedio'),'dias_estimados':first(cols,'dias_estimados','dias'),'fecha_estimada':first(cols,'fecha_estimada','fecha'),'estado_operativo':first(cols,'estado_operativo','estado'),'prioridad':first(cols,'prioridad'),'accion_sugerida':first(cols,'accion_sugerida','accion')}
    return {}

def all_rows(t,limit=3000):
    if not table_exists(t): return []
    try: return q(f'SELECT * FROM {t} LIMIT {int(limit)}')
    except: return []

def ensure_schema():
    if engine is None: return
    pk=pk_sql()
    q('''CREATE TABLE IF NOT EXISTS usuarios (id SERIAL PRIMARY KEY, usuario TEXT UNIQUE, nombre TEXT, password_hash TEXT, rol TEXT, activo BOOLEAN DEFAULT TRUE, creado TIMESTAMP)''', fetch=False)
    q('''CREATE TABLE IF NOT EXISTS maestro_equipos (codigo TEXT PRIMARY KEY, tipo_equipo TEXT, familia TEXT, marca TEXT, modelo TEXT, ano TEXT, ubicacion TEXT, responsable TEXT, lectura_actual TEXT, unidad TEXT, proxima_pm TEXT, estado TEXT, descripcion TEXT)''', fetch=False)
    q(f'''CREATE TABLE IF NOT EXISTS lecturas (id {pk}, fecha DATE, codigo TEXT, horometro NUMERIC, kilometraje NUMERIC, obra_ubicacion TEXT, responsable TEXT, observacion TEXT)''', fetch=False)
    q(f'''CREATE TABLE IF NOT EXISTS mantenciones (id {pk}, fecha DATE, codigo TEXT, tipo_mantencion TEXT, lectura NUMERIC, espm TEXT, folio TEXT, lugar TEXT, proveedor TEXT, costo_mantencion_clp TEXT, estado TEXT)''', fetch=False)
    q(f'''CREATE TABLE IF NOT EXISTS ot (id {pk}, fecha DATE, ot TEXT, codigo TEXT, tipo TEXT, lectura TEXT, descripcion TEXT, responsable TEXT, estado TEXT, costo TEXT)''', fetch=False)
    q(f'''CREATE TABLE IF NOT EXISTS compras (id {pk}, fecha DATE, oc TEXT, codigo TEXT, descripcion TEXT, proveedor TEXT, costo_pm_clp TEXT, regla TEXT, estado_oc TEXT)''', fetch=False)
    q(f'''CREATE TABLE IF NOT EXISTS bodega (id {pk}, folio TEXT, fecha DATE, equipo TEXT, envio TEXT, persona_que_retiro TEXT, destino TEXT, comentario TEXT, codigo TEXT)''', fetch=False)
    q(f'''CREATE TABLE IF NOT EXISTS plan_mantenciones (id {pk}, codigo TEXT, tipo_equipo TEXT, familia TEXT, control TEXT, lectura_actual TEXT, proxima_lectura_objetivo TEXT, promedio_diario TEXT, dias_estimados TEXT, fecha_estimada TEXT, estado_operativo TEXT, costo_total_pm TEXT, prioridad TEXT, accion_sugerida TEXT)''', fetch=False)
    q(f'''CREATE TABLE IF NOT EXISTS plan_90_dias (id {pk}, codigo TEXT, tipo_equipo TEXT, familia TEXT, control TEXT, lectura_actual TEXT, proxima_lectura_objetivo TEXT, promedio_diario TEXT, dias_estimados TEXT, fecha_estimada TEXT, estado_operativo TEXT, prioridad TEXT, accion_sugerida TEXT)''', fetch=False)
    q(f'''CREATE TABLE IF NOT EXISTS calendario_pm (id {pk}, fecha DATE, codigo TEXT, tipo_pm TEXT, descripcion TEXT, responsable TEXT, estado TEXT, observacion TEXT)''', fetch=False)
    for t,cols in {'usuarios':[('nombre','TEXT'),('password_hash','TEXT'),('rol','TEXT'),('activo','BOOLEAN'),('creado','TIMESTAMP')],'maestro_equipos':[('tipo_equipo','TEXT'),('familia','TEXT'),('marca','TEXT'),('modelo','TEXT'),('ano','TEXT'),('ubicacion','TEXT'),('responsable','TEXT'),('lectura_actual','TEXT'),('unidad','TEXT'),('proxima_pm','TEXT'),('estado','TEXT'),('descripcion','TEXT')],'lecturas':[('fecha','DATE'),('codigo','TEXT'),('horometro','NUMERIC'),('kilometraje','NUMERIC'),('obra_ubicacion','TEXT'),('responsable','TEXT'),('observacion','TEXT')],'mantenciones':[('fecha','DATE'),('codigo','TEXT'),('tipo_mantencion','TEXT'),('lectura','NUMERIC'),('espm','TEXT'),('folio','TEXT'),('lugar','TEXT'),('proveedor','TEXT'),('costo_mantencion_clp','TEXT'),('estado','TEXT')],'ot':[('fecha','DATE'),('ot','TEXT'),('codigo','TEXT'),('tipo','TEXT'),('lectura','TEXT'),('descripcion','TEXT'),('responsable','TEXT'),('estado','TEXT'),('costo','TEXT')],'compras':[('fecha','DATE'),('oc','TEXT'),('codigo','TEXT'),('descripcion','TEXT'),('proveedor','TEXT'),('costo_pm_clp','TEXT'),('estado_oc','TEXT')],'bodega':[('folio','TEXT'),('fecha','DATE'),('equipo','TEXT'),('envio','TEXT'),('persona_que_retiro','TEXT'),('destino','TEXT'),('comentario','TEXT'),('codigo','TEXT')]}.items():
        if table_exists(t):
            for c,d in cols: add_col(t,c,d)
    try:
        if not q('SELECT * FROM usuarios LIMIT 1'):
            q('INSERT INTO usuarios (usuario,nombre,password_hash,rol,activo,creado) VALUES (:u,:n,:p,:r,:a,:c)', {'u':'admin','n':'Administrador','p':generate_password_hash('admin123'),'r':'admin','a':True,'c':datetime.now()}, fetch=False)
    except: pass

def ultima_lectura_real(codigo):
    codigo=(codigo or '').upper(); cand=[]
    if table_exists('lecturas'):
        m=map_cols('lecturas')
        for r in all_rows('lecturas',10000):
            if val(r,m.get('codigo')).upper()==codigo:
                fecha=val(r,m.get('fecha')); ubi=norm_ubic(val(r,m.get('obra_ubicacion')))
                h=r.get(m.get('horometro')) if m.get('horometro') else None; k=r.get(m.get('kilometraje')) if m.get('kilometraje') else None
                if h not in [None,'']: cand.append({'fecha':fecha,'valor':num(h),'unidad':'HORAS','origen':'Lectura','ubicacion':ubi})
                if k not in [None,'']: cand.append({'fecha':fecha,'valor':num(k),'unidad':'KM','origen':'Lectura','ubicacion':ubi})
    if table_exists('mantenciones'):
        m=map_cols('mantenciones')
        for r in all_rows('mantenciones',10000):
            if val(r,m.get('codigo')).upper()==codigo:
                lec=r.get(m.get('lectura')) if m.get('lectura') else None
                if lec not in [None,'']: cand.append({'fecha':val(r,m.get('fecha')),'valor':num(lec),'unidad':'','origen':'Mantención','ubicacion':norm_ubic(val(r,m.get('lugar')))})
    if not cand: return {'fecha':'','valor':'','unidad':'','origen':'','ubicacion':''}
    def fk(x):
        try: return datetime.strptime(str(x.get('fecha') or '1900-01-01')[:10],'%Y-%m-%d')
        except: return datetime(1900,1,1)
    return sorted(cand,key=fk,reverse=True)[0]

def get_equipos():
    table='maestro_equipos' if table_exists('maestro_equipos') else 'equipos'; m=map_cols(table); data=[]
    for r in all_rows(table,10000):
        cod=val(r,m.get('codigo'))
        if not cod: continue
        ult=ultima_lectura_real(cod); lec=val(r,m.get('lectura_actual')); uni=val(r,m.get('unidad'))
        if ult.get('valor') not in [None,'']:
            lec=str(int(ult['valor']) if float(ult['valor']).is_integer() else ult['valor']); uni=ult.get('unidad') or uni
        ubi=norm_ubic(val(r,m.get('ubicacion')))
        if ult.get('ubicacion'): ubi=ult['ubicacion']
        data.append({'codigo':cod,'tipo_equipo':val(r,m.get('tipo_equipo')),'familia':val(r,m.get('familia')),'marca':val(r,m.get('marca')),'modelo':val(r,m.get('modelo')),'ano':val(r,m.get('ano')),'ubicacion':ubi,'responsable':val(r,m.get('responsable')),'lectura_actual':lec,'unidad':uni,'proxima_pm':val(r,m.get('proxima_pm')),'estado':val(r,m.get('estado')) or 'Sin estado','descripcion':val(r,m.get('descripcion')),'ultima_fecha':ult.get('fecha',''),'ultima_origen':ult.get('origen','')})
    return sorted(data,key=lambda x:str(x['codigo']))
def get_equipo(c):
    c=(c or '').upper(); return next((e for e in get_equipos() if e['codigo'].upper()==c),None)
def is_atrasado(e): return 'ATRAS' in e.get('estado','').upper() or 'VENC' in e.get('estado','').upper()
def is_proximo(e): return any(x in e.get('estado','').upper() for x in ['PROX','RECIBIR','PROCESO'])
def is_aldia(e): return 'AL D' in e.get('estado','').upper()
def is_fuera(e): return any(x in e.get('estado','').upper() for x in ['FUERA','TALLER'])

def historial_data(codigo):
    c=(codigo or '').upper(); out=[]
    for table,origen in [('lecturas','Lectura'),('mantenciones','Mantención'),('ot','OT'),('compras','Compra'),('bodega','Bodega')]:
        if not table_exists(table): continue
        m=map_cols(table)
        for r in all_rows(table,10000):
            code=val(r,m.get('codigo')) or val(r,m.get('equipo'))
            if str(code).upper()==c:
                out.append({'fecha':val(r,m.get('fecha')),'origen':origen,'detalle':val(r,m.get('descripcion')) or val(r,m.get('tipo_mantencion')) or val(r,m.get('espm')) or val(r,m.get('comentario')) or val(r,m.get('destino')),'lectura':val(r,m.get('lectura')) or val(r,m.get('ot')) or val(r,m.get('oc')) or val(r,m.get('folio')),'estado':val(r,m.get('estado')) or val(r,m.get('estado_oc'))})
    return sorted(out,key=lambda x:str(x.get('fecha') or ''),reverse=True)

def plan_rows(src='plan_mantenciones'):
    data=[]; table=src if table_exists(src) else 'plan_mantenciones'
    if table_exists(table):
        m=map_cols(table)
        for r in all_rows(table,10000):
            cod=val(r,m.get('codigo'))
            if not cod: continue
            e=get_equipo(cod) or {}; dias=num(r.get(m.get('dias_estimados')) if m.get('dias_estimados') else 0); estado=val(r,m.get('estado_operativo')) or e.get('estado','')
            data.append({'codigo':cod,'tipo_equipo':val(r,m.get('tipo_equipo')) or e.get('tipo_equipo',''),'ubicacion':e.get('ubicacion',''),'control':val(r,m.get('control')) or e.get('unidad',''),'lectura_actual':val(r,m.get('lectura_actual')) or e.get('lectura_actual',''),'proxima':val(r,m.get('proxima_lectura_objetivo')) or e.get('proxima_pm',''),'dias':dias,'fecha':val(r,m.get('fecha_estimada')),'estado':estado or ('ATRASADA' if dias<0 else 'PRÓXIMA' if dias<=15 else 'AL DÍA'),'accion':val(r,m.get('accion_sugerida'))})
    if not data:
        for e in get_equipos(): data.append({'codigo':e['codigo'],'tipo_equipo':e['tipo_equipo'],'ubicacion':e['ubicacion'],'control':e['unidad'],'lectura_actual':e['lectura_actual'],'proxima':e['proxima_pm'],'dias':-1 if is_atrasado(e) else 10 if is_proximo(e) else 60,'fecha':'','estado':e['estado'],'accion':'Programar/revisar'})
    return sorted(data,key=lambda x:(x['dias'],x['codigo']))

def login_required(fn):
    @wraps(fn)
    def w(*a,**k):
        if not session.get('user'): return redirect(url_for('login'))
        return fn(*a,**k)
    return w
def admin_required(fn):
    @wraps(fn)
    def w(*a,**k):
        if session.get('rol')!='admin': return redirect(url_for('dashboard'))
        return fn(*a,**k)
    return w

def equipo_datalist(): return '<datalist id="equiposList">'+''.join(f"<option value='{e['codigo']}'>{e['codigo']} - {e['tipo_equipo']}</option>" for e in get_equipos())+'</datalist>'
def badge(s):
    u=str(s or '').upper(); cls='badge bad' if 'ATRAS' in u or 'VENC' in u else 'badge warn' if any(x in u for x in ['PROX','RECIBIR','PROCESO']) else 'badge off' if any(x in u for x in ['FUERA','TALLER']) else 'badge'
    return f"<span class='{cls}'>{safe(s)}</span>"
def estado_select(name='estado',current=''): return f"<select name='{name}'>"+''.join(f"<option value='{e}' {'selected' if e==current else ''}>{e}</option>" for e in ESTADOS)+'</select>'
def ubicacion_select(name='ubicacion',current=''): return f"<select name='{name}'>"+''.join(f"<option value='{u}' {'selected' if u==norm_ubic(current) else ''}>{u}</option>" for u in UBICACIONES)+'</select>'
def tipo_pm_select(name='tipo',current=''): return f"<select name='{name}'>"+''.join(f"<option value='{t}' {'selected' if t==current else ''}>{t}</option>" for t in TIPOS_PM)+'</select>'
def kpi_card(cls,icon,title,value,sub): return f"<div class='kpi-card {cls}'><div class='kpi-icon'>{icon}</div><div><small>{title}</small><b>{value}</b><span>{sub}</span></div></div>"
def machine_svg(): return "<svg viewBox='0 0 140 80'><circle cx='42' cy='61' r='10' fill='#263238'/><circle cx='92' cy='61' r='10' fill='#263238'/><rect x='30' y='38' width='70' height='20' rx='4' fill='#e6a400'/><rect x='70' y='24' width='30' height='20' rx='3' fill='#ffbf1b'/><path d='M96 37 L123 21 L127 27 L103 47' stroke='#d59100' stroke-width='7' fill='none'/><rect x='113' y='44' width='18' height='10' rx='2' fill='#333'/></svg>"
def vertical_bars(data,label='Equipos'):
    if not data: return "<div class='empty'>Sin datos</div>"
    mx=max(v for _,v in data) or 1
    return "<div class='vertical-bars'>"+''.join(f"<div class='vbar-item'><strong>{v}</strong><div class='vbar' style='height:{max(12,int(v/mx*145))}px'></div><small>{k}</small></div>" for k,v in data)+f"</div><div class='legend-dot red'>{label}</div>"
def page(title,body,extra=''):
    user=session.get('user','Administrador')
    return render_template_string(f"""<!doctype html><html lang='es'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>{title}</title><link rel='icon' href='/static/img/favicon.svg' type='image/svg+xml'><link rel='stylesheet' href='/static/css/styles.css'></head><body><header class='topbar'><a class='logo' href='/'>DEMOTRON</a><nav class='nav'><a href='/' class='active'>▦ Dashboard</a><a href='/equipos'>⚙ Equipos</a><a href='/ot'>🛠 OT</a><a href='/lecturas'>▤ Lecturas</a><a href='/compras'>🛒 Compras</a><a href='/bodega'>⌂ Bodega</a><a href='/reportes'>▣ Reportes</a><a href='/admin'>⚙ Configuración</a></nav><form class='topsearch' action='/ficha'><input name='codigo' list='equiposList' placeholder='Buscar...'>{equipo_datalist()}</form><div class='bell'>🔔<span>5</span></div><div class='userchip'>{user} ▾</div></header>{body}<footer class='footer'><b>DEMOTRON</b> CMMS <span>Versión 2.0.0</span></footer><script src='/static/js/cmms.js'></script>{extra}</body></html>""")

@app.before_request
def before_any():
    if request.endpoint!='static':
        try: ensure_schema()
        except Exception: pass
@app.route("/login", methods=["GET", "POST"])
def login():
    error = ""
    if request.method == "POST":
        u = (request.form.get("usuario") or "").strip()
        p = request.form.get("password") or ""

        # Acceso de respaldo DEMOTRON: siempre disponible aunque la tabla usuarios esté vieja o dañada.
        if u == "admin" and p == "admin123":
            session["user"] = "admin"
            session["rol"] = "admin"
            return redirect(url_for("dashboard"))

        try:
            row = q("SELECT * FROM usuarios WHERE usuario=:u LIMIT 1", {"u": u})
            if row:
                r = row[0]
                activo = r.get("activo", True)
                if str(activo).lower() in ["false", "0", "no"]:
                    error = "Usuario desactivado"
                else:
                    password_hash = r.get("password_hash") or ""
                    password_plano = r.get("password") or r.get("clave") or ""
                    ok_hash = False
                    if password_hash:
                        try:
                            ok_hash = check_password_hash(password_hash, p)
                        except Exception:
                            ok_hash = False
                    ok_plano = bool(password_plano and str(password_plano) == p)
                    if ok_hash or ok_plano:
                        session["user"] = r.get("usuario") or u
                        session["rol"] = r.get("rol") or "usuario"
                        return redirect(url_for("dashboard"))
                    else:
                        error = "Usuario o contraseña incorrectos"
            else:
                error = "Usuario o contraseña incorrectos"
        except Exception:
            error = "No se pudo validar usuarios. Use admin / admin123."

    return render_template_string(f"""<!doctype html><html><head>{CSS}</head><body class="login-body"><form class="login-card" method="post"><h1>DEMOTRON</h1><p>CMMS ERP</p><input name="usuario" value="admin"><input type="password" name="password" value="admin123">{'<div class="error">'+error+'</div>' if error else ''}<button>Entrar</button></form></body></html>""")


@app.route('/logout')
def logout(): session.clear(); return redirect(url_for('login'))
@app.route('/api/version')
def version(): return jsonify({'version':'DEMOTRON_VISUAL_ADMIN_FINAL','status':'ok'})

@app.route("/admin/forzar-admin")
def forzar_admin():
    try:
        ensure_schema()
        q("DELETE FROM usuarios WHERE usuario='admin'", fetch=False)
        q("""INSERT INTO usuarios (usuario,nombre,password_hash,rol,activo,creado)
             VALUES (:usuario,:nombre,:password_hash,:rol,:activo,:creado)""",
          {"usuario": "admin", "nombre": "Administrador", "password_hash": generate_password_hash("admin123"), "rol": "admin", "activo": True, "creado": datetime.now()}, fetch=False)
        return jsonify({"ok": True, "usuario": "admin", "password": "admin123"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route('/api/status')
def status(): return jsonify({'database':dialect(),'status':'ok','tables':tables()})
@app.route('/')
@login_required
def dashboard():
    equipos=get_equipos(); total=len(equipos); atr=[e for e in equipos if is_atrasado(e)]; atrasados=len(atr); prox=sum(1 for e in equipos if is_proximo(e)); aldia=sum(1 for e in equipos if is_aldia(e)); fuera=sum(1 for e in equipos if is_fuera(e)); control=round(aldia/total*100) if total else 0
    ot_count=len(all_rows('ot',10000)) if table_exists('ot') else 0; compras=all_rows('compras',10000) if table_exists('compras') else []; cm=map_cols('compras'); compra_monto=sum(num(r.get(cm.get('costo_pm_clp'))) for r in compras) if cm else 0
    ubic={}
    for e in atr: ubic[e['ubicacion'] or 'Sin ubicación']=ubic.get(e['ubicacion'] or 'Sin ubicación',0)+1
    ubic_data=sorted(ubic.items(),key=lambda x:x[1],reverse=True)[:5] or [('Sin datos',0)]
    p1=int(aldia/max(total,1)*100); p2=int(prox/max(total,1)*100); p3=int(atrasados/max(total,1)*100)
    donut=f"background:conic-gradient(#45b86a 0 {p1}%,#f2b718 {p1}% {p1+p2}%,#f23c45 {p1+p2}% {p1+p2+p3}%,#9aa1af {p1+p2+p3}% 100%)"
    rows=''.join(f"<tr><td><a href='/equipo/{e['codigo']}'><b>{e['codigo']}</b></a></td><td>{e['descripcion'] or e['tipo_equipo']}</td><td>{e['ubicacion']}</td><td>{e['lectura_actual']}</td><td>{e['proxima_pm']}</td><td class='redtxt'>-</td><td>{badge(e['estado'])}</td><td><a class='pillbtn' href='/ot/nueva?codigo={e['codigo']}'>Crear OT</a></td></tr>" for e in atr[:10]) or "<tr><td colspan='8'>No hay equipos atrasados.</td></tr>"
    quick=''.join(f"<a class='equip-card {'danger' if is_atrasado(e) else ''}' href='/equipo/{e['codigo']}'><span class='dot {'red' if is_atrasado(e) else 'green' if is_aldia(e) else 'yellow'}'></span><div class='mini-img'>{machine_svg()}</div><h4>{e['codigo']}</h4><p>{e['tipo_equipo']}<br>{e['marca']} {e['modelo']}</p><small>Lectura: {e['lectura_actual']}</small></a>" for e in atr[:12])
    body=f"""<main class='page'><section class='kpis'>{kpi_card('red','!', 'ATRASADOS', atrasados, f'{round(atrasados/max(total,1)*100,1)}% del total')}{kpi_card('yellow','◷','PRÓXIMOS',prox,f'{round(prox/max(total,1)*100,1)}% del total')}{kpi_card('green','✓','CONTROLADO REAL',f'{control}%',f'{aldia} de {total} equipos')}{kpi_card('blue','▣','OT ABIERTAS',ot_count,'+ esta semana')}{kpi_card('purple','🛒','COMPRAS EN PROCESO',len(compras),clp(compra_monto))}{kpi_card('teal','$','COSTO MENSUAL',clp(compra_monto),'Compras PM')}</section><section class='dashboard-grid'><div class='panel'><h3>ESTADO GENERAL DE LA FLOTA</h3><div class='donut-wrap'><div class='donut' style='{donut}'><span>{total}<small>Equipos</small></span></div><div class='legend-list'><p><i></i>Controlado<b>{aldia}</b></p><p><i></i>Próximos<b>{prox}</b></p><p><i></i>Atrasados<b>{atrasados}</b></p><p><i></i>Fuera servicio<b>{fuera}</b></p></div></div></div><div class='panel'><h3>ATRASADOS POR UBICACIÓN</h3>{vertical_bars(ubic_data,'Equipos atrasados')}</div><div class='panel'><h3>GESTIÓN (OT Y COMPRAS)</h3>{vertical_bars([('Esta semana',ot_count),('Semana anterior',max(0,ot_count-5)),('Este mes',ot_count+len(compras)),('Mes anterior',max(0,ot_count-2))],'OT / Compras')}</div></section><section class='lower-grid'><div class='panel wide'><div class='section-head'><h3>EQUIPOS ATRASADOS (CRÍTICOS)</h3><a href='/equipos'>Ver todos</a></div><table><thead><tr><th>Equipo</th><th>Descripción</th><th>Ubicación</th><th>Última lectura</th><th>Próxima PM</th><th>Margen</th><th>Estado</th><th>Acción</th></tr></thead><tbody>{rows}</tbody></table></div><div class='panel'><div class='section-head'><h3>ACTIVIDAD RECIENTE</h3><a href='/reportes'>Ver todo</a></div><div class='activity'><p>🟩 OT creada recientemente</p><p>🟦 Lectura registrada</p><p>🛒 Compra PM registrada</p><p>📦 Repuesto ingresado a bodega</p></div></div></section><section class='panel'><h3>EQUIPOS ATRASADOS (VISTA RÁPIDA)</h3><div class='cards-scroll'>{quick}</div></section></main>"""
    return page('Dashboard',body)

@app.route('/admin',methods=['GET','POST'])
@login_required
@admin_required
def admin():
    msg=''
    if request.method=='POST' and request.form.get('action')=='crear_usuario':
        data={'usuario':clean(request.form.get('usuario')),'nombre':clean(request.form.get('nombre')) or clean(request.form.get('usuario')),'password_hash':generate_password_hash(clean(request.form.get('password')) or '123456'),'rol':clean(request.form.get('rol')) or 'consulta','activo':True,'creado':datetime.now()}
        try: q('INSERT INTO usuarios (usuario,nombre,password_hash,rol,activo,creado) VALUES (:usuario,:nombre,:password_hash,:rol,:activo,:creado)',data,fetch=False); msg='Usuario creado correctamente.'
        except Exception as e: msg=f'No se pudo crear usuario: {e}'
    users=q('SELECT usuario,nombre,rol,activo,creado FROM usuarios ORDER BY id DESC LIMIT 100') if table_exists('usuarios') else []
    rows=''.join(f"<tr><td>{u['usuario']}</td><td>{u['nombre']}</td><td>{u['rol']}</td><td>{u['activo']}</td><td>{safe(u['creado'])}</td></tr>" for u in users)
    body=f"""<main class='page'><section class='panel'><h2>Configuración / Administración</h2><p>{msg}</p></section><section class='admin-grid'><div class='panel'><h3>Subir Excel</h3><form class='admin-form' method='post' action='/admin/upload-excel' enctype='multipart/form-data'><input type='file' name='archivo' accept='.xlsx'><button>Subir e importar</button></form><p>Hojas reconocidas: Maestro_Equipos, Lecturas, Mantenciones, Compras PM, Bodega, Plan_Mantenciones, Plan_90_Dias.</p></div><div class='panel'><h3>Crear usuario</h3><form method='post' class='admin-form'><input type='hidden' name='action' value='crear_usuario'><input name='usuario' placeholder='Usuario'><input name='nombre' placeholder='Nombre'><input name='password' placeholder='Contraseña'><select name='rol'><option value='admin'>Administrador</option><option value='mantencion'>Mantención</option><option value='consulta'>Consulta</option></select><button>Crear usuario</button></form></div></section><section class='panel'><h3>Usuarios</h3><table><thead><tr><th>Usuario</th><th>Nombre</th><th>Rol</th><th>Activo</th><th>Creado</th></tr></thead><tbody>{rows}</tbody></table></section></main>"""
    return page('Administración',body)

@app.route('/admin/upload-excel',methods=['POST'])
@login_required
@admin_required
def upload_excel():
    f=request.files.get('archivo')
    if not f: return redirect(url_for('admin'))
    try:
        from openpyxl import load_workbook
        wb=load_workbook(filename=BytesIO(f.read()),data_only=True)
        mapping={'maestro_equipos':'maestro_equipos','maestro equipos':'maestro_equipos','lecturas':'lecturas','mantenciones':'mantenciones','compras pm':'compras','compras':'compras','bodega':'bodega','plan_mantenciones':'plan_mantenciones','plan mantenciones':'plan_mantenciones','plan_90_dias':'plan_90_dias','plan 90 dias':'plan_90_dias'}
        imported={}
        for ws in wb.worksheets:
            table=mapping.get(norm_key(ws.title).replace('_',' ')) or mapping.get(norm_key(ws.title))
            if not table: continue
            rows=list(ws.iter_rows(values_only=True));
            if not rows: continue
            hi=0
            for i,row in enumerate(rows[:20]):
                vals=[norm_key(x) for x in row if x is not None]
                if 'codigo' in vals or 'fecha' in vals or 'folio' in vals: hi=i; break
            headers=[]; seen={}
            for i,x in enumerate(rows[hi]):
                h=norm_key(x) or f'col_{i}'
                if h in seen: seen[h]+=1; h=f'{h}_{seen[h]}'
                else: seen[h]=0
                headers.append(h); add_col(table,h,'TEXT')
            count=0
            for row in rows[hi+1:]:
                if not row or all(v is None for v in row): continue
                data={headers[i]:safe(row[i]) if i<len(row) else '' for i in range(len(headers))}
                q(f"INSERT INTO {table} ({','.join(headers)}) VALUES ({','.join(':'+h for h in headers)})",data,fetch=False); count+=1
            imported[table]=imported.get(table,0)+count
        return jsonify({'ok':True,'imported':imported})
    except Exception as e: return jsonify({'ok':False,'error':str(e)}),500

@app.route('/equipos',methods=['GET','POST'])
@login_required
def equipos():
    if request.method=='POST':
        data={k:clean(request.form.get(k)) for k in ['tipo_equipo','familia','marca','modelo','ano','responsable','lectura_actual','unidad','proxima_pm','estado','descripcion']}; data['codigo']=up(request.form.get('codigo')); data['ubicacion']=norm_ubic(request.form.get('ubicacion'))
        q('DELETE FROM maestro_equipos WHERE UPPER(codigo)=UPPER(:codigo)',{'codigo':data['codigo']},fetch=False); q('INSERT INTO maestro_equipos (codigo,tipo_equipo,familia,marca,modelo,ano,ubicacion,responsable,lectura_actual,unidad,proxima_pm,estado,descripcion) VALUES (:codigo,:tipo_equipo,:familia,:marca,:modelo,:ano,:ubicacion,:responsable,:lectura_actual,:unidad,:proxima_pm,:estado,:descripcion)',data,fetch=False); return redirect(url_for('equipos'))
    data=get_equipos(); rows=''.join(f"<tr><td><a href='/equipo/{e['codigo']}'><b>{e['codigo']}</b></a></td><td>{e['tipo_equipo']}</td><td>{e['familia']}</td><td>{e['marca']}</td><td>{e['modelo']}</td><td>{e['ubicacion']}</td><td>{e['lectura_actual']}</td><td>{e['unidad']}</td><td>{badge(e['estado'])}</td></tr>" for e in data)
    form=f"<form class='form-grid' method='post'><input name='codigo' list='equiposList' placeholder='Código'>{equipo_datalist()}<input name='tipo_equipo' placeholder='Tipo equipo'><input name='familia' placeholder='Familia'><input name='marca' placeholder='Marca'><input name='modelo' placeholder='Modelo'><input name='ano' placeholder='Año'>{ubicacion_select()}<input name='lectura_actual' placeholder='Lectura actual'><select name='unidad'><option>HORAS</option><option>KM</option></select>{estado_select()}<button>Guardar equipo</button></form>"
    return page('Equipos',f"<main class='page'><div class='panel'><h2>Equipos</h2>{form}<table><thead><tr><th>Código</th><th>Tipo</th><th>Familia</th><th>Marca</th><th>Modelo</th><th>Ubicación</th><th>Lectura</th><th>Unidad</th><th>Estado</th></tr></thead><tbody>{rows}</tbody></table></div></main>",f"<script>window.EQUIPOS={json.dumps(data,ensure_ascii=False)};</script>")
@app.route('/ficha')
@login_required
def ficha_redirect():
    c=request.args.get('codigo',''); return redirect(url_for('equipo_ficha',codigo=c)) if c else redirect(url_for('dashboard'))
@app.route('/equipo/<codigo>')
@login_required
def equipo_ficha(codigo):
    e=get_equipo(codigo)
    if not e: return page('No encontrado',f"<main class='page'><div class='panel'>No existe {codigo}</div></main>")
    hist=historial_data(e['codigo']); rows=''.join(f"<tr><td>{h['fecha']}</td><td>{h['origen']}</td><td>{h['detalle']}</td><td>{h['lectura']}</td><td>{h['estado']}</td></tr>" for h in hist[:100])
    return page('Ficha',f"<main class='page'><section class='panel'><h2>{e['codigo']} · {e['tipo_equipo']}</h2><p>{e['marca']} {e['modelo']} · {e['ubicacion']} · {badge(e['estado'])}</p><p><a class='pillbtn' href='/lecturas?codigo={e['codigo']}'>Agregar lectura</a> <a class='pillbtn' href='/ot/nueva?codigo={e['codigo']}'>Crear OT</a></p></section><section class='panel'><h3>Historial técnico</h3><table><thead><tr><th>Fecha</th><th>Origen</th><th>Detalle</th><th>Lectura/Folio</th><th>Estado</th></tr></thead><tbody>{rows}</tbody></table></section></main>")
@app.route('/lecturas',methods=['GET','POST'])
@login_required
def lecturas():
    if request.method=='POST':
        data={'fecha':clean(request.form.get('fecha')),'codigo':up(request.form.get('codigo')),'horometro':number(request.form.get('horometro')),'kilometraje':number(request.form.get('kilometraje')),'obra_ubicacion':norm_ubic(request.form.get('obra_ubicacion')),'responsable':clean(request.form.get('responsable')),'observacion':clean(request.form.get('observacion'))}; q('INSERT INTO lecturas (fecha,codigo,horometro,kilometraje,obra_ubicacion,responsable,observacion) VALUES (:fecha,:codigo,:horometro,:kilometraje,:obra_ubicacion,:responsable,:observacion)',data,fetch=False); return redirect(url_for('lecturas'))
    c=request.args.get('codigo',''); m=map_cols('lecturas'); rows=''.join(f"<tr><td>{val(r,m.get('fecha'))}</td><td><a href='/equipo/{val(r,m.get('codigo'))}'><b>{val(r,m.get('codigo'))}</b></a></td><td>{safe(r.get(m.get('horometro')) if m.get('horometro') else '')}</td><td>{safe(r.get(m.get('kilometraje')) if m.get('kilometraje') else '')}</td><td>{norm_ubic(val(r,m.get('obra_ubicacion')))}</td><td>{val(r,m.get('responsable'))}</td></tr>" for r in all_rows('lecturas'))
    form=f"<form class='form-grid' method='post'><input name='codigo' list='equiposList' value='{c}' placeholder='Código'>{equipo_datalist()}<input type='date' name='fecha'><input type='number' step='any' name='horometro' placeholder='Horómetro'><input type='number' step='any' name='kilometraje' placeholder='Kilometraje'>{ubicacion_select('obra_ubicacion')}<input name='responsable' placeholder='Responsable'><input name='observacion' placeholder='Observación'><button>Guardar lectura</button></form>"
    return page('Lecturas',f"<main class='page'><div class='panel'><h2>Lecturas</h2>{form}<table><thead><tr><th>Fecha</th><th>Código</th><th>Horómetro</th><th>Kilometraje</th><th>Ubicación</th><th>Responsable</th></tr></thead><tbody>{rows}</tbody></table></div></main>")
@app.route('/mantenciones',methods=['GET','POST'])
@login_required
def mantenciones():
    if request.method=='POST':
        data={'fecha':clean(request.form.get('fecha')),'codigo':up(request.form.get('codigo')),'tipo_mantencion':clean(request.form.get('tipo_mantencion')),'lectura':number(request.form.get('lectura')),'espm':clean(request.form.get('espm')),'folio':clean(request.form.get('folio')) or f"OT-{datetime.now().strftime('%Y%m%d%H%M%S')}",'lugar':norm_ubic(request.form.get('lugar')),'proveedor':clean(request.form.get('proveedor')),'costo_mantencion_clp':clean(request.form.get('costo_mantencion_clp')),'estado':clean(request.form.get('estado')) or 'EN PROCESO'}; q('INSERT INTO mantenciones (fecha,codigo,tipo_mantencion,lectura,espm,folio,lugar,proveedor,costo_mantencion_clp,estado) VALUES (:fecha,:codigo,:tipo_mantencion,:lectura,:espm,:folio,:lugar,:proveedor,:costo_mantencion_clp,:estado)',data,fetch=False); return redirect(url_for('mantenciones'))
    m=map_cols('mantenciones'); rows=''.join(f"<tr><td>{val(r,m.get('fecha'))}</td><td>{val(r,m.get('codigo'))}</td><td>{val(r,m.get('tipo_mantencion'))}</td><td>{val(r,m.get('lectura'))}</td><td>{val(r,m.get('folio'))}</td><td>{val(r,m.get('proveedor'))}</td><td>{badge(val(r,m.get('estado')))}</td></tr>" for r in all_rows('mantenciones'))
    form=f"<form class='form-grid' method='post'><input name='codigo' list='equiposList' placeholder='Código'>{equipo_datalist()}<input type='date' name='fecha'><select name='tipo_mantencion'>{''.join(f'<option>{t}</option>' for t in TIPOS_PM)}</select><input type='number' step='any' name='lectura' placeholder='Lectura'><input name='espm' placeholder='Descripción'><input name='folio' placeholder='Folio / OT'>{ubicacion_select('lugar')}<input name='proveedor' placeholder='Proveedor'>{estado_select()}<button>Guardar mantención</button></form>"
    return page('Mantenciones',f"<main class='page'><div class='panel'><h2>Mantenciones</h2>{form}<table><thead><tr><th>Fecha</th><th>Código</th><th>Tipo</th><th>Lectura</th><th>Folio</th><th>Proveedor</th><th>Estado</th></tr></thead><tbody>{rows}</tbody></table></div></main>")
@app.route('/ot')
@login_required
def ot():
    m=map_cols('ot'); rows=''.join(f"<tr><td>{val(r,m.get('fecha'))}</td><td><a href='/ot/{val(r,m.get('id'))}'><b>{val(r,m.get('ot'))}</b></a></td><td>{val(r,m.get('codigo'))}</td><td>{val(r,m.get('tipo'))}</td><td>{val(r,m.get('descripcion'))}</td><td>{badge(val(r,m.get('estado')))}</td><td><a class='pillbtn' href='/ot/{val(r,m.get('id'))}/pdf'>PDF</a></td></tr>" for r in all_rows('ot'))
    return page('OT',f"<main class='page'><div class='panel'><div class='section-head'><h2>Órdenes de Trabajo</h2><a class='pillbtn' href='/ot/nueva'>Nueva OT</a></div><table><thead><tr><th>Fecha</th><th>OT</th><th>Equipo</th><th>Tipo</th><th>Descripción</th><th>Estado</th><th>PDF</th></tr></thead><tbody>{rows}</tbody></table></div></main>")
@app.route('/ot/nueva',methods=['GET','POST'])
@login_required
def ot_nueva():
    c=request.args.get('codigo','')
    if request.method=='POST':
        data={'fecha':clean(request.form.get('fecha')) or date.today().isoformat(),'ot':clean(request.form.get('ot')) or f"OT-{datetime.now().strftime('%Y%m%d%H%M%S')}",'codigo':up(request.form.get('codigo')),'tipo':clean(request.form.get('tipo')),'lectura':clean(request.form.get('lectura')),'descripcion':clean(request.form.get('descripcion')),'responsable':clean(request.form.get('responsable')),'estado':clean(request.form.get('estado')) or 'EN PROCESO','costo':None}; q('INSERT INTO ot (fecha,ot,codigo,tipo,lectura,descripcion,responsable,estado,costo) VALUES (:fecha,:ot,:codigo,:tipo,:lectura,:descripcion,:responsable,:estado,:costo)',data,fetch=False); return redirect(url_for('ot'))
    e=get_equipo(c) if c else None; lectura=f"{e['lectura_actual']} {e['unidad']}" if e else ''
    form=f"<form class='form-grid' method='post'><input type='date' name='fecha' value='{date.today().isoformat()}'><input name='ot' placeholder='N° OT automático'><input name='codigo' list='equiposList' value='{c}' placeholder='Código'>{equipo_datalist()}{tipo_pm_select()}<input name='lectura' value='{lectura}' placeholder='Lectura actual'><textarea name='descripcion' placeholder='Descripción'></textarea><input name='responsable' placeholder='Responsable'>{estado_select(current='EN PROCESO')}<button>Crear OT</button></form>"
    return page('Nueva OT',f"<main class='page'><div class='panel'><h2>Nueva Orden de Trabajo</h2>{form}</div></main>")
@app.route('/ot/<int:ot_id>')
@login_required
def ot_detalle(ot_id): return page('Detalle OT',f"<main class='page'><div class='panel'><h2>OT {ot_id}</h2><a class='pillbtn' href='/ot/{ot_id}/pdf'>Descargar PDF</a></div></main>")
@app.route('/ot/<int:ot_id>/pdf')
@login_required
def ot_pdf(ot_id):
    row=q('SELECT * FROM ot WHERE id=:id',{'id':ot_id}) if col_exists('ot','id') else []
    if not row: return 'OT no encontrada',404
    r=row[0]; m=map_cols('ot'); path=f'/tmp/OT_{ot_id}.pdf'
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.pdfgen import canvas
        c=canvas.Canvas(path,pagesize=letter); w,h=letter; c.setFillColor(colors.HexColor('#073a7a')); c.rect(0,h-2.2*cm,w,2.2*cm,stroke=0,fill=1); c.setFillColor(colors.white); c.setFont('Helvetica-Bold',18); c.drawString(1.2*cm,h-1.35*cm,'DEMOTRON - ORDEN DE TRABAJO'); c.setFillColor(colors.black); c.setFont('Helvetica',10); c.drawString(1.2*cm,h-3*cm,f"OT: {val(r,m.get('ot'))}  Equipo: {val(r,m.get('codigo'))}  Tipo: {val(r,m.get('tipo'))}"); c.drawString(1.2*cm,h-4*cm,f"Descripción: {val(r,m.get('descripcion'))}"[:110]); c.save()
    except Exception as e: return str(e),500
    return send_file(path,as_attachment=True,download_name=f'OT_{ot_id}.pdf')
@app.route('/compras')
@login_required
def compras():
    m=map_cols('compras'); rows=''.join(f"<tr><td>{val(r,m.get('fecha'))}</td><td>{val(r,m.get('oc'))}</td><td>{val(r,m.get('codigo'))}</td><td>{val(r,m.get('descripcion'))}</td><td>{val(r,m.get('proveedor'))}</td><td>{clp(r.get(m.get('costo_pm_clp')) if m.get('costo_pm_clp') else '')}</td><td>{badge(val(r,m.get('estado_oc')))}</td></tr>" for r in all_rows('compras'))
    return page('Compras',f"<main class='page'><div class='panel'><h2>Compras</h2><table><thead><tr><th>Fecha</th><th>OC</th><th>Código</th><th>Descripción</th><th>Proveedor</th><th>Monto</th><th>Estado</th></tr></thead><tbody>{rows}</tbody></table></div></main>")
@app.route('/bodega')
@login_required
def bodega():
    m=map_cols('bodega'); rows=''.join(f"<tr><td>{val(r,m.get('folio'))}</td><td>{val(r,m.get('fecha'))}</td><td>{val(r,m.get('equipo'))}</td><td>{val(r,m.get('envio'))}</td><td>{val(r,m.get('persona_que_retiro'))}</td><td>{val(r,m.get('destino'))}</td><td>{val(r,m.get('comentario'))}</td><td>{val(r,m.get('codigo'))}</td></tr>" for r in all_rows('bodega'))
    return page('Bodega',f"<main class='page'><div class='panel'><h2>Bodega</h2><table><thead><tr><th>Folio</th><th>Fecha</th><th>Equipo</th><th>Envío</th><th>Retira</th><th>Destino</th><th>Comentario</th><th>Código</th></tr></thead><tbody>{rows}</tbody></table></div></main>")
@app.route('/reportes')
@login_required
def reportes(): return page('Reportes',"<main class='page'><div class='panel'><h2>Reportes</h2><p>Módulo preparado.</p></div></main>")
def plan_page(title,src):
    rows=''.join(f"<tr><td><a href='/equipo/{r['codigo']}'><b>{r['codigo']}</b></a></td><td>{r['tipo_equipo']}</td><td>{r['ubicacion']}</td><td>{r['control']}</td><td>{r['lectura_actual']}</td><td>{r['proxima']}</td><td>{int(r['dias'])}</td><td>{r['fecha']}</td><td>{badge(r['estado'])}</td><td>{r['accion']}</td></tr>" for r in plan_rows(src))
    return page(title,f"<main class='page'><div class='panel'><h2>{title}</h2><table><thead><tr><th>Equipo</th><th>Tipo</th><th>Ubicación</th><th>Control</th><th>Lectura</th><th>Próxima</th><th>Días</th><th>Fecha</th><th>Estado</th><th>Acción</th></tr></thead><tbody>{rows}</tbody></table></div></main>")
@app.route('/planificacion')
@login_required
def planificacion(): return plan_page('Plan Mantenciones','plan_mantenciones')
@app.route('/plan-90-dias')
@login_required
def plan90(): return plan_page('Plan 90 días','plan_90_dias')
@app.route('/calendario')
@login_required
def calendario(): return page('Calendario',"<main class='page'><div class='panel'><h2>Calendario</h2><p>Calendario PM manual preparado.</p></div></main>")
@app.route('/backlog')
@login_required
def backlog(): return plan_page('Backlog PM','plan_mantenciones')
if __name__=='__main__': app.run(host='0.0.0.0', port=int(os.getenv('PORT',8080)))
